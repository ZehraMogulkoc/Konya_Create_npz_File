from openpyxl import load_workbook
import pandas as pd
# Dosya yolu
dosya_yolu = r"C:\\\\Users\\\\Lenovo\\Downloads\\\\Konya Kavşak Sayım Verileri\\\\Konya Kavşak Sayım Verileri\\Hakimiyet Kavşağı\\Ocak\\10_Ocak\\Tüm yönler.xlsx"

pd.read_excel(dosya_yolu)
# Dosyayı yükle
dosya = load_workbook(dosya_yolu)

# Dosyadaki sayfaları listele
sayfa_isimleri = dosya.sheetnames
print("Sayfa isimleri:", sayfa_isimleri)

# İlk sayfayı al
ilk_sayfa = dosya[sayfa_isimleri[0]]

# Örnek olarak, ilk 5 satırı yazdır
for satir in ilk_sayfa.iter_rows(min_row=1):
    for hucre in satir:
        print(hucre.value, end="\t")
    print()

# Dosyayı kapat
dosya.close()
